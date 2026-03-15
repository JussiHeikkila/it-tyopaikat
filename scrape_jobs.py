#!/usr/bin/env python3
"""
IT-työpaikkojen päivittäinen seuranta.

Kaksi datanlähdettä:
  1. Duunitori (automaattinen, ei vaadi kirjautumista)
  2. Joblistings API (manuaalinen, vaatii Bearer-tokenin)

Käyttö:
  python3 scrape_jobs.py                  # Duunitori (automaattinen)
  python3 scrape_jobs.py --token XXXX     # Joblistings API
"""

import argparse
import json
import re
import sys
import time
from datetime import datetime, date
from pathlib import Path
from urllib.parse import quote_plus

import requests
from bs4 import BeautifulSoup

DATA_DIR = Path(__file__).parent / "data"
DATA_DIR.mkdir(exist_ok=True)

HISTORY_FILE = DATA_DIR / "history.json"

# === Kategorioiden luokittelusäännöt ===
# Käytetään sekä Duunitori-haussa että API-ilmoitusten luokittelussa.
# "keywords" = sanat joilla ilmoitus matchataan (job_title + required_skills)
CATEGORIES = {
    "Ohjelmistokehitys": {
        "duunitori": {"terms": ["ohjelmistokehitys", "+software +developer", "+software +engineer"], "mode": "sum"},
        "keywords": ["software developer", "software engineer", "ohjelmistokehit", "sovelluskehit",
                      "ohjelmistosuunnit", "ohjelmoija", "lead developer", "senior developer",
                      "developer,", "kehittäjä", "kehittaja", "graafikko"],
    },
    "Fullstack": {
        "duunitori": {"terms": ["fullstack", "full-stack"], "mode": "best"},
        "keywords": ["fullstack", "full-stack", "full stack"],
    },
    "Frontend": {
        "duunitori": {"terms": ["frontend", "front-end"], "mode": "best"},
        "keywords": ["frontend", "front-end", "front end", "react developer", "vue developer", "angular developer"],
    },
    "Backend": {
        "duunitori": {"terms": ["backend", "back-end"], "mode": "best"},
        "keywords": ["backend", "back-end", "back end"],
    },
    "Embedded / Sulautetut": {
        "duunitori": {"terms": ["+embedded +software", "+embedded +developer", "+sulautetut +järjestelmät"], "mode": "sum"},
        "keywords": ["embedded", "sulautettu", "sulautetut"],
    },
    "Mobiilisovelluskehitys": {
        "duunitori": {"terms": ["+mobile +developer", "+ios +developer", "+android +developer", "mobiilikehittäjä"], "mode": "sum"},
        "keywords": ["mobile developer", "ios developer", "android developer", "mobiili", "react native",
                      "flutter", "mobile engineer", "android engineer", "ios engineer", "kotlin"],
    },
    "Design (UX/UI)": {
        "duunitori": {"terms": ["+ux +designer", "+ui +designer"], "mode": "sum"},
        "keywords": ["ux designer", "ui designer", "ux/ui", "ux ", "ui "],
    },
    "Testaus / QA": {
        "duunitori": {"terms": ["+test +engineer", "+qa +engineer", "testausinsinööri"], "mode": "sum"},
        "keywords": ["test engineer", "qa engineer", "quality assurance", "testaus", "testaaja", "sdet",
                      "test lead", "validation engineer", "test automation", "robot framework", "playwright"],
    },
    "Data & BI": {
        "duunitori": {"terms": ["+data +analyst", "+power +bi", "+business +intelligence", "+data +analyytikko"], "mode": "sum"},
        "keywords": ["data analyst", "data-analyy", "liiketoiminta-analyy", "business analyst",
                      "power bi", "tableau", "analytics engineer", "analytiikka",
                      "tietovarasto", "data steward", "data governance", "raportointi"],
    },
    "Data Scientist": {
        "duunitori": {"terms": ["+data +scientist"], "mode": "best"},
        "keywords": ["data scientist", "data science"],
    },
    "Data Engineer": {
        "duunitori": {"terms": ["+data +engineer"], "mode": "best"},
        "keywords": ["data engineer"],
    },
    "Data Architect": {
        "duunitori": {"terms": ["+data +architect"], "mode": "best"},
        "keywords": ["data architect"],
    },
    "DevOps / SRE": {
        "duunitori": {"terms": ["+devops +engineer", "+site +reliability"], "mode": "sum"},
        "keywords": ["devops", "sre", "site reliability", "platform engineer", "infrastructure as code"],
    },
    "Tietoturva / IAM": {
        "duunitori": {"terms": ["+cybersecurity", "+tietoturva", "+information +security"], "mode": "sum"},
        "keywords": ["cybersecurity", "tietoturva", "information security", "security engineer",
                      "iam ", "identity management", "access management", "penetration test"],
    },
    "IT-projektinhallinta": {
        "duunitori": {"terms": ["+it +projektipäällikkö", "+scrum +master", "+agile +coach"], "mode": "sum"},
        "keywords": ["projektipäällikkö", "project manager", "scrum master", "agile coach",
                      "delivery manager"],
    },
    "Järjestelmäasiantuntija": {
        "duunitori": {"terms": ["järjestelmäasiantuntija", "+system +specialist", "+it +specialist"], "mode": "sum"},
        "keywords": ["järjestelmäasiantuntija", "system specialist", "system engineer",
                      "it operations", "it-tuki", "it specialist"],
    },
    "SAP / ERP": {
        "duunitori": {"terms": ["+sap +consultant", "+erp +developer", "+dynamics +365"], "mode": "sum"},
        "keywords": ["sap ", "erp", "dynamics 365", "dynamics365", "salesforce",
                      "successfactors", "guidewire"],
    },
    "SW Architect": {
        "duunitori": {"terms": ["+software +architect", "+solution +architect"], "mode": "sum"},
        "keywords": ["software architect", "ohjelmistoarkkitehti", "solution architect",
                      "solution designer", "solution engineer"],
    },
    "Cloud Engineer": {
        "duunitori": {"terms": ["+cloud +engineer"], "mode": "best"},
        "keywords": ["cloud engineer", "aws kehit", "azure engineer"],
    },
    "Cloud Architect": {
        "duunitori": {"terms": ["+cloud +architect"], "mode": "best"},
        "keywords": ["cloud architect"],
    },
    "GenAI Engineer": {
        "duunitori": {"terms": ["+genai +engineer", "+generative +ai +engineer"], "mode": "best"},
        "keywords": ["genai engineer", "generative ai engineer"],
    },
    "AI Engineer": {
        "duunitori": {"terms": ["+ai +engineer", "+machine +learning +engineer", "+ml +engineer"], "mode": "sum"},
        "keywords": ["ai engineer", "machine learning engineer", "ml engineer",
                      "ai performance", "agentic", "computer vision", "perception software"],
    },
    "GenAI Architect": {
        "duunitori": {"terms": ["+genai +architect", "+ai +architect"], "mode": "best"},
        "keywords": ["genai architect", "ai architect"],
    },
}

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
    "Accept-Language": "fi-FI,fi;q=0.9,en;q=0.8",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

SESSION = requests.Session()
SESSION.headers.update(HEADERS)

JOBLISTINGS_API = "https://joblistings-agent-lnojw7sokq-lz.a.run.app"


# =====================================================================
# Lähde 1: Duunitori (scraping)
# =====================================================================

def fetch_duunitori_count(search_term: str) -> int:
    url = f"https://duunitori.fi/tyopaikat?haku={quote_plus(search_term)}&alue=suomi"
    try:
        resp = SESSION.get(url, timeout=15)
        resp.raise_for_status()
        match = re.search(r'"results_count"\s*:\s*"(\d+)"', resp.text)
        return int(match.group(1)) if match else 0
    except Exception as e:
        print(f"  [VIRHE] Duunitori '{search_term}': {e}", file=sys.stderr)
        return 0


def get_duunitori_count(terms: list, mode: str) -> int:
    counts = []
    for term in terms:
        counts.append(fetch_duunitori_count(term))
        time.sleep(0.5)
    return sum(counts) if mode == "sum" else (max(counts) if counts else 0)


def scrape_duunitori() -> dict:
    today = date.today().isoformat()
    results = {"date": today, "source": "duunitori", "categories": {}, "timestamp": datetime.now().isoformat()}
    total = len(CATEGORIES)
    for i, (cat, cfg) in enumerate(CATEGORIES.items(), 1):
        d = cfg["duunitori"]
        print(f"[{i}/{total}] {cat}...", end=" ", flush=True)
        count = get_duunitori_count(d["terms"], d["mode"])
        print(count)
        results["categories"][cat] = {"count": count}
    return results


# =====================================================================
# Lähde 2: Joblistings API
# =====================================================================

def fetch_all_joblistings(token: str) -> list:
    """Hae kaikki ilmoitukset sivutetusti."""
    headers = {"Authorization": f"Bearer {token}", "Accept": "application/json"}
    all_listings = []
    skip = 0
    limit = 100

    print("Haetaan ilmoituksia Joblistings API:sta...", flush=True)

    while True:
        url = f"{JOBLISTINGS_API}/api/v1/job-listings/?skip={skip}&limit={limit}"
        resp = requests.get(url, headers=headers, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        listings = data.get("data", [])
        total = data.get("total", 0)
        all_listings.extend(listings)
        print(f"  Haettu {len(all_listings)}/{total}...", flush=True)

        if len(all_listings) >= total or not listings:
            break
        skip += limit
        time.sleep(0.3)

    print(f"  Yhteensä {len(all_listings)} ilmoitusta")
    return all_listings


def load_excel_listings(filepath: str) -> list:
    """Lue Excel-tiedosto ja muunna listaksi dictejä (sama muoto kuin API)."""
    import openpyxl
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    listings = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        item = dict(zip(headers, row))
        listings.append({
            "job_title": item.get("Title") or "",
            "required_skills": [s.strip() for s in (item.get("Skills") or "").split(",") if s.strip()],
            "company_name": item.get("Company") or "",
            "location": item.get("Location") or "",
        })
    return listings


def classify_listing(listing: dict) -> list:
    """Luokittele yksittäinen ilmoitus kategorioihin job_title + required_skills perusteella."""
    title = (listing.get("job_title") or "").lower()
    skills = [s.lower() for s in (listing.get("required_skills") or [])]
    searchable = title + " " + " ".join(skills)

    matched = []
    for cat, cfg in CATEGORIES.items():
        for kw in cfg["keywords"]:
            if kw.lower() in searchable:
                matched.append(cat)
                break
    return matched


def classify_listings(listings: list) -> dict:
    """Luokittele valmiiksi haetut ilmoitukset."""
    today = date.today().isoformat()
    results = {
        "date": today,
        "source": "joblistings-api",
        "categories": {},
        "timestamp": datetime.now().isoformat(),
        "total_listings": len(listings),
    }

    counts = {cat: 0 for cat in CATEGORIES}
    unclassified = 0
    for listing in listings:
        cats = classify_listing(listing)
        if cats:
            for cat in cats:
                counts[cat] += 1
        else:
            unclassified += 1

    for cat, count in counts.items():
        results["categories"][cat] = {"count": count}

    print(f"Luokiteltu {len(listings)} ilmoitusta:")
    sorted_cats = sorted(counts.items(), key=lambda x: x[1], reverse=True)
    for cat, count in sorted_cats:
        print(f"  {cat:30s} {count:5d}")
    print(f"\n  {'Luokittelematta':30s} {unclassified:5d}")
    print(f"  {'API ilmoituksia yhteensä':30s} {len(listings):5d}")

    return results


def scrape_joblistings(token: str) -> dict:
    """Hae ja luokittele kaikki Joblistings API:n ilmoitukset."""
    listings = fetch_all_joblistings(token)

    today = date.today().isoformat()
    results = {
        "date": today,
        "source": "joblistings-api",
        "categories": {},
        "timestamp": datetime.now().isoformat(),
        "total_listings": len(listings),
    }

    # Laske per kategoria
    counts = {cat: 0 for cat in CATEGORIES}
    unclassified = 0
    for listing in listings:
        cats = classify_listing(listing)
        if cats:
            for cat in cats:
                counts[cat] += 1
        else:
            unclassified += 1

    for cat, count in counts.items():
        results["categories"][cat] = {"count": count}

    # Tulosta yhteenveto
    print(f"\nLuokiteltu {len(listings)} ilmoitusta:")
    sorted_cats = sorted(counts.items(), key=lambda x: x[1], reverse=True)
    for cat, count in sorted_cats:
        print(f"  {cat:30s} {count:5d}")
    print(f"\n  {'Luokittelematta':30s} {unclassified:5d}")
    print(f"  {'API ilmoituksia yhteensä':30s} {len(listings):5d}")

    return results


# =====================================================================
# HTML-generointi ja pääohjelma
# =====================================================================

def load_history() -> list:
    if HISTORY_FILE.exists():
        with open(HISTORY_FILE) as f:
            return json.load(f)
    return []


def save_history(history: list):
    with open(HISTORY_FILE, "w") as f:
        json.dump(history, f, indent=2, ensure_ascii=False)


def _build_source_table(counts, categories, colors):
    """Rakenna yhden lähteen taulukkorivit."""
    rows = []
    for cat in categories:
        rows.append({"category": cat, "count": counts.get(cat, 0)})
    rows.sort(key=lambda r: r["count"], reverse=True)
    max_count = max((r["count"] for r in rows), default=1) or 1
    html = ""
    for row in rows:
        if row["count"] == 0:
            continue
        pct = (row["count"] / max_count) * 100
        ci = list(CATEGORIES.keys()).index(row["category"]) % len(colors)
        color = colors[ci]
        html += f"""                    <tr>
                        <td>{row["category"]}</td>
                        <td class="num count-col">{row["count"]}</td>
                        <td class="bar-cell"><div class="bar-bg"><div class="bar-fill" style="width:{pct:.0f}%;background:{color}"></div></div></td>
                    </tr>
"""
    return html


def generate_html(history: list):
    html_file = Path(__file__).parent / "index.html"
    categories = list(CATEGORIES.keys())

    # Erota lähteet: viimeisin data kummastakin
    latest_duunitori = None
    latest_joblistings = None
    for h in reversed(history):
        src = h.get("source", "duunitori")
        if src == "duunitori" and not latest_duunitori:
            latest_duunitori = h
        elif src == "joblistings-api" and not latest_joblistings:
            latest_joblistings = h
        if latest_duunitori and latest_joblistings:
            break

    colors = [
        "#3B82F6", "#EF4444", "#10B981", "#F59E0B", "#8B5CF6",
        "#EC4899", "#06B6D4", "#84CC16", "#F97316", "#6366F1",
        "#14B8A6", "#E11D48", "#0EA5E9", "#A855F7", "#D946EF",
        "#22C55E", "#FF6B6B", "#0891B2", "#65A30D", "#C026D3",
        "#DC2626", "#059669", "#7C3AED",
    ]

    def get_counts(data):
        if not data:
            return {}
        return {cat: data["categories"].get(cat, {}).get("count", 0) for cat in categories}

    duunitori_counts = get_counts(latest_duunitori)
    joblistings_counts = get_counts(latest_joblistings)

    duunitori_total = sum(duunitori_counts.values()) if duunitori_counts else 0
    joblistings_total = sum(joblistings_counts.values()) if joblistings_counts else 0
    api_all = latest_joblistings.get("total_listings", 0) if latest_joblistings else 0

    duunitori_time = latest_duunitori["timestamp"][:16].replace("T", " ") if latest_duunitori else "-"
    joblistings_time = latest_joblistings["timestamp"][:16].replace("T", " ") if latest_joblistings else "-"

    duunitori_rows = _build_source_table(duunitori_counts, categories, colors)
    joblistings_rows = _build_source_table(joblistings_counts, categories, colors)

    # Chart: Duunitori-historia
    duunitori_history = [h for h in history if h.get("source", "duunitori") == "duunitori"][-30:]
    d_dates = [h["date"] for h in duunitori_history]
    d_datasets = []
    for i, cat in enumerate(categories):
        values = [h["categories"].get(cat, {}).get("count", 0) for h in duunitori_history]
        d_datasets.append({
            "label": cat, "data": values,
            "borderColor": colors[i % len(colors)],
            "backgroundColor": colors[i % len(colors)] + "20",
            "tension": 0.3, "borderWidth": 2,
        })

    html = f"""<!DOCTYPE html>
<html lang="fi">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>IT-työpaikat Suomessa</title>
    <script src="https://cdn.jsdelivr.net/npm/chart.js@4"></script>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', system-ui, sans-serif;
            background: #0f172a; color: #e2e8f0; min-height: 100vh;
        }}
        .header {{
            background: linear-gradient(135deg, #1e293b, #334155);
            padding: 2rem; border-bottom: 1px solid #475569;
        }}
        .header h1 {{ font-size: 1.75rem; font-weight: 700; color: #f8fafc; }}
        .header .subtitle {{ color: #94a3b8; margin-top: 0.25rem; font-size: 0.9rem; }}
        .header .sources {{ color: #64748b; font-size: 0.8rem; margin-top: 0.5rem; }}
        .header .sources a {{ color: #60a5fa; text-decoration: none; }}
        .badge {{ display: inline-block; padding: 0.15rem 0.5rem; border-radius: 0.25rem; font-size: 0.75rem; font-weight: 600; }}
        .badge-api {{ background: #065f46; color: #6ee7b7; }}
        .badge-scrape {{ background: #1e3a5f; color: #93c5fd; }}
        .kpi-bar {{
            display: flex; gap: 1.5rem; padding: 1.5rem 2rem;
            background: #1e293b; border-bottom: 1px solid #334155; flex-wrap: wrap;
        }}
        .kpi {{ background: #334155; padding: 1rem 1.5rem; border-radius: 0.75rem; min-width: 160px; }}
        .kpi .label {{ color: #94a3b8; font-size: 0.8rem; text-transform: uppercase; letter-spacing: 0.05em; }}
        .kpi .value {{ font-size: 1.75rem; font-weight: 700; color: #f8fafc; margin-top: 0.25rem; }}
        .kpi .sub {{ color: #64748b; font-size: 0.8rem; margin-top: 0.25rem; }}
        .container {{ padding: 2rem; max-width: 1400px; margin: 0 auto; }}
        .card {{
            background: #1e293b; border: 1px solid #334155;
            border-radius: 0.75rem; padding: 1.5rem; margin-bottom: 2rem;
        }}
        .card h2 {{ font-size: 1.1rem; color: #f8fafc; margin-bottom: 1rem; padding-bottom: 0.75rem; border-bottom: 1px solid #334155; }}
        /* Tabs */
        .tabs {{
            display: flex; gap: 0; margin-bottom: 2rem;
        }}
        .tab {{
            padding: 0.75rem 1.5rem; cursor: pointer; font-weight: 600; font-size: 0.95rem;
            border: 1px solid #334155; border-bottom: none;
            border-radius: 0.5rem 0.5rem 0 0; background: #0f172a; color: #64748b;
            transition: all 0.2s;
        }}
        .tab:hover {{ color: #94a3b8; background: #1e293b; }}
        .tab.active {{ background: #1e293b; color: #f8fafc; border-color: #475569; }}
        .tab-content {{ display: none; }}
        .tab-content.active {{ display: block; }}
        /* Dual panel layout */
        .dual-panels {{
            display: grid; grid-template-columns: 1fr 1fr; gap: 2rem;
        }}
        @media (max-width: 1024px) {{
            .dual-panels {{ grid-template-columns: 1fr; }}
        }}
        table {{ width: 100%; border-collapse: collapse; font-size: 0.9rem; }}
        th {{
            text-align: left; padding: 0.75rem 1rem; background: #334155;
            color: #94a3b8; font-weight: 600; font-size: 0.8rem;
            text-transform: uppercase; letter-spacing: 0.05em;
        }}
        th:first-child {{ border-radius: 0.5rem 0 0 0.5rem; }}
        th:last-child {{ border-radius: 0 0.5rem 0.5rem 0; }}
        td {{ padding: 0.75rem 1rem; border-bottom: 1px solid #0f172a; }}
        tr:nth-child(even) {{ background: #1e293b; }}
        tr:nth-child(odd) {{ background: #162032; }}
        tr:hover {{ background: #334155; transition: background 0.15s; }}
        .num {{ text-align: right; font-variant-numeric: tabular-nums; }}
        .count-col {{ font-weight: 700; color: #f8fafc; font-size: 1rem; }}
        .bar-cell {{ width: 40%; }}
        .bar-bg {{ background: #334155; border-radius: 0.25rem; height: 1.25rem; overflow: hidden; }}
        .bar-fill {{ height: 100%; border-radius: 0.25rem; transition: width 0.3s; }}
        .chart-container {{ position: relative; height: 420px; }}
        .footer {{ text-align: center; padding: 2rem; color: #475569; font-size: 0.8rem; }}
        .note {{ color: #64748b; font-size: 0.8rem; margin-top: 0.75rem; font-style: italic; }}
        @media (max-width: 768px) {{
            .kpi-bar {{ flex-direction: column; }}
            .container {{ padding: 1rem; }}
            .bar-cell {{ display: none; }}
            table {{ font-size: 0.8rem; }}
            td, th {{ padding: 0.5rem; }}
        }}
    </style>
</head>
<body>
    <div class="header">
        <h1>IT-työpaikat Suomessa</h1>
        <div class="subtitle">Avointen IT-työpaikkojen seuranta kahdesta lähteestä</div>
        <div class="sources">
            <a href="https://duunitori.fi" target="_blank">Duunitori</a> <span class="badge badge-scrape">SCRAPE</span>
            &nbsp;&amp;&nbsp;
            <a href="https://joblistings.aiexp.fi" target="_blank">Joblistings / Ailandai</a> <span class="badge badge-api">API</span>
        </div>
    </div>

    <div class="kpi-bar">
        <div class="kpi">
            <div class="label">Duunitori</div>
            <div class="value" style="color:#93c5fd">{duunitori_total:,}</div>
            <div class="sub">luokitellut ({duunitori_time})</div>
        </div>
        <div class="kpi">
            <div class="label">Joblistings</div>
            <div class="value" style="color:#6ee7b7">{joblistings_total:,}</div>
            <div class="sub">luokitellut ({joblistings_time})</div>
        </div>
        <div class="kpi">
            <div class="label">Joblistings kaikki</div>
            <div class="value">{api_all:,}</div>
            <div class="sub">ilmoitukset yhteensä</div>
        </div>
        <div class="kpi">
            <div class="label">Kategorioita</div>
            <div class="value">{len(categories)}</div>
            <div class="sub">seurannassa</div>
        </div>
    </div>

    <div class="container">
        <div class="tabs">
            <div class="tab active" onclick="switchTab('side-by-side')">Rinnakkain</div>
            <div class="tab" onclick="switchTab('duunitori')">Duunitori</div>
            <div class="tab" onclick="switchTab('joblistings')">Joblistings</div>
        </div>

        <!-- Rinnakkain -->
        <div id="tab-side-by-side" class="tab-content active">
            <div class="dual-panels">
                <div class="card">
                    <h2>Duunitori <span class="badge badge-scrape">SCRAPE</span></h2>
                    <table>
                        <thead><tr><th>Tehtävä</th><th class="num">Avoimia</th><th class="bar-cell"></th></tr></thead>
                        <tbody>
{duunitori_rows}                        </tbody>
                    </table>
                    <div class="note">Aggregoi TE-palvelut, yrityssivustot ym. Luvut = hakutulosten määrä.</div>
                </div>
                <div class="card">
                    <h2>Joblistings <span class="badge badge-api">API</span></h2>
                    <table>
                        <thead><tr><th>Tehtävä</th><th class="num">Avoimia</th><th class="bar-cell"></th></tr></thead>
                        <tbody>
{joblistings_rows}                        </tbody>
                    </table>
                    <div class="note">AI-luokitellut ilmoitukset joblistings.aiexp.fi:stä ({api_all} ilmoitusta yhteensä).</div>
                </div>
            </div>
        </div>

        <!-- Duunitori -->
        <div id="tab-duunitori" class="tab-content">
            <div class="card">
                <h2>Duunitori — avoimet työpaikat <span class="badge badge-scrape">SCRAPE</span></h2>
                <table>
                    <thead><tr><th>Tehtävä</th><th class="num">Avoimia</th><th class="bar-cell"></th></tr></thead>
                    <tbody>
{duunitori_rows}                    </tbody>
                </table>
                <div class="note">Duunitori aggregoi useita lähteitä (TE-palvelut, yrityssivustot ym.). Päivitetty {duunitori_time}.</div>
            </div>
            <div class="card">
                <h2>Duunitori — kehitys (viimeiset 30 päivää)</h2>
                <div class="chart-container">
                    <canvas id="trendChart"></canvas>
                </div>
            </div>
        </div>

        <!-- Joblistings -->
        <div id="tab-joblistings" class="tab-content">
            <div class="card">
                <h2>Joblistings — avoimet työpaikat <span class="badge badge-api">API</span></h2>
                <table>
                    <thead><tr><th>Tehtävä</th><th class="num">Avoimia</th><th class="bar-cell"></th></tr></thead>
                    <tbody>
{joblistings_rows}                    </tbody>
                </table>
                <div class="note">{api_all} ilmoitusta yhteensä, joista {joblistings_total} luokiteltu IT-kategorioihin. Päivitetty {joblistings_time}.</div>
            </div>
        </div>
    </div>

    <div class="footer">
        Data: Duunitori &amp; Joblistings API (joblistings.aiexp.fi)
    </div>

    <script>
        function switchTab(name) {{
            document.querySelectorAll('.tab-content').forEach(el => el.classList.remove('active'));
            document.querySelectorAll('.tab').forEach(el => el.classList.remove('active'));
            document.getElementById('tab-' + name).classList.add('active');
            event.target.classList.add('active');
            // Init chart when Duunitori tab shown
            if (name === 'duunitori' && !window.chartInit) {{
                initChart();
            }}
        }}

        function initChart() {{
            window.chartInit = true;
            const ctx = document.getElementById('trendChart').getContext('2d');
            new Chart(ctx, {{
                type: 'line',
                data: {{
                    labels: {json.dumps(d_dates)},
                    datasets: {json.dumps(d_datasets, ensure_ascii=False)}
                }},
                options: {{
                    responsive: true,
                    maintainAspectRatio: false,
                    interaction: {{ mode: 'index', intersect: false }},
                    plugins: {{
                        legend: {{
                            position: 'bottom',
                            labels: {{ color: '#94a3b8', boxWidth: 12, padding: 15, font: {{ size: 11 }} }}
                        }},
                        tooltip: {{
                            backgroundColor: '#1e293b', borderColor: '#475569', borderWidth: 1,
                            titleColor: '#f8fafc', bodyColor: '#e2e8f0',
                        }}
                    }},
                    scales: {{
                        x: {{ ticks: {{ color: '#64748b' }}, grid: {{ color: '#1e293b' }} }},
                        y: {{ ticks: {{ color: '#64748b' }}, grid: {{ color: '#1e293b' }}, beginAtZero: true }}
                    }}
                }}
            }});
        }}
    </script>
</body>
</html>"""

    with open(html_file, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"\nHTML-raportti: {html_file}")


def main():
    parser = argparse.ArgumentParser(description="IT-työpaikkojen seuranta")
    parser.add_argument("--token", help="Joblistings API Bearer-token (manuaalinen ajo)")
    parser.add_argument("--file", help="JSON-tiedosto Joblistings API:n exportista (manuaalinen ajo)")
    args = parser.parse_args()

    print(f"=== IT-työpaikkojen seuranta {date.today().isoformat()} ===\n")

    if args.file:
        if args.file.endswith((".xlsx", ".xls")):
            print(f"Lähde: Excel-tiedosto {args.file}\n")
            listings = load_excel_listings(args.file)
        else:
            print(f"Lähde: JSON-tiedosto {args.file}\n")
            with open(args.file, encoding="utf-8") as f:
                listings = json.load(f)
        today_results = classify_listings(listings)
    elif args.token:
        print("Lähde: Joblistings API\n")
        today_results = scrape_joblistings(args.token)
    else:
        print("Lähde: Duunitori\n")
        today_results = scrape_duunitori()

    history = load_history()
    today_str = date.today().isoformat()
    history = [h for h in history if h["date"] != today_str]
    history.append(today_results)

    save_history(history)
    print(f"\nData: {HISTORY_FILE}")

    generate_html(history)

    total = sum(d["count"] for d in today_results["categories"].values())
    print(f"\nYHTEENSÄ: {total}")


if __name__ == "__main__":
    main()
